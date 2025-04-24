import os
import re
import json
import pandas as pd
import subprocess
from openpyxl import load_workbook

# ----------------------------------------
def parse_value(val):
    if isinstance(val, str):
        v = val.strip()
        if v.lower() == "true":   return True
        if v.lower() == "false":  return False
        try:    return int(v)
        except:
            try:    return float(v)
            except: return v
    return val

# ----------------------------------------
def set_nested_value(d, key_str, value):
    parts = key_str.split(".")
    cur = d
    for i, part in enumerate(parts):
        m = re.match(r"(\w+)(?:\[(\d+)\])?", part)
        if not m:
            continue
        key, idx = m.group(1), m.group(2)
        if i == len(parts) - 1:
            if idx is not None:
                j = int(idx)
                cur.setdefault(key, [])
                while len(cur[key]) <= j:
                    cur[key].append(None)
                cur[key][j] = value
            else:
                cur[key] = value
        else:
            if idx is not None:
                j = int(idx)
                cur.setdefault(key, [])
                while len(cur[key]) <= j:
                    cur[key].append({})
                if cur[key][j] is None:
                    cur[key][j] = {}
                cur = cur[key][j]
            else:
                cur.setdefault(key, {})
                cur = cur[key]

# ----------------------------------------
def format_primitive(v):
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    # エスケープ: "${" → "$${"
    safe = v.replace("${", "$${")
    return f'"{safe}"'

# ----------------------------------------
def dict_to_hcl(d, indent=0):
    ind = "  " * indent
    lines = []
    for k, v in d.items():
        if isinstance(v, dict):
            lines.append(f"{ind}{k} = {{")
            lines.append(dict_to_hcl(v, indent + 1))
            lines.append(f"{ind}}}")
        elif isinstance(v, list):
            lines.append(f"{ind}{k} = {list_to_hcl(v, indent)}")
        else:
            lines.append(f"{ind}{k} = {format_primitive(v)}")
    return "\n".join(lines)

# ----------------------------------------
def list_to_hcl(lst, indent=0):
    ind = "  " * indent
    inner = "  " * (indent + 1)
    lines = ["["]
    for it in lst:
        if isinstance(it, dict):
            lines.append(inner + "{")
            lines.append(dict_to_hcl(it, indent + 2))
            lines.append(inner + "},")
        elif isinstance(it, list):
            lines.append(inner + list_to_hcl(it, indent + 1) + ",")
        else:
            lines.append(inner + format_primitive(it) + ",")
    lines.append(ind + "]")
    return "\n".join(lines)

# ----------------------------------------
def sheet_to_df(wb, name):
    ws = wb[name]
    vals = ws.values
    try:
        header = next(vals)
    except StopIteration:
        return pd.DataFrame()
    return pd.DataFrame(vals, columns=header)

# ----------------------------------------
def main():
    parameter_dir = "parameter"
    modules_root  = "modules"
    os.makedirs(modules_root, exist_ok=True)

    # 1) シート名マッピング読み込み
    mapping = {}
    cfg_path = "sheetname_config.json"
    if os.path.exists(cfg_path):
        with open(cfg_path, "r", encoding="utf-8") as f:
            mapping = json.load(f)

    tfvars_data = {}
    env_value = None

    # 2) common.xlsx を読み込み、共通キーを tfvars_data にセット
    common_path = os.path.join(parameter_dir, "common.xlsx")
    if os.path.exists(common_path):
        wb_c = load_workbook(common_path, data_only=True, read_only=True)
        first_sheet = wb_c.sheetnames[0]
        df_c = sheet_to_df(wb_c, first_sheet)
        if {"key", "value"}.issubset(df_c.columns):
            for _, row in df_c.iterrows():
                k = str(row["key"]).strip()
                v = parse_value(row["value"])
                tfvars_data[k] = v
                # env キーを検出
                if k == "env":
                    env_value = str(v)
        else:
            print("common.xlsx: 列名は 'key','value' をご利用ください。")

    # env が取得できなければエラー終了
    if not env_value:
        raise RuntimeError("common.xlsx に 'env' キーが定義されていません。")

    # 3) 各サービスExcel の処理
    for fn in os.listdir(parameter_dir):
        if not fn.lower().endswith(".xlsx") or fn.lower() == "common.xlsx":
            continue
        wb_path    = os.path.join(parameter_dir, fn)
        service    = os.path.splitext(fn)[0]  # e.g. "s3", "iam"
        module_dir = os.path.join(modules_root, service)
        os.makedirs(module_dir, exist_ok=True)

        try:
            wb = load_workbook(wb_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"{fn} 読み込み失敗: {e}")
            continue

        for sheet in wb.sheetnames:
            df = sheet_to_df(wb, sheet)
            if df.empty or "resource-name" not in df.columns:
                continue

            actual_key = mapping.get(sheet, sheet)
            top_key    = actual_key + "_list"
            tfvars_data.setdefault(top_key, {})

            for res in df["resource-name"].dropna().unique():
                name = str(res).strip()
                grp  = df[df["resource-name"].astype(str).str.strip() == name]
                tfvars_data[top_key].setdefault(name, {})

                # policy JSON を modules/<service>/<name>.json に出力
                if grp["arguments"].astype(str).str.strip().eq("policy").any():
                    for _, r in grp.iterrows():
                        if str(r["arguments"]).strip() != "policy":
                            continue
                        raw = r.get("value", "")
                        try:
                            obj = json.loads(raw)
                        except Exception as ex:
                            print(f"{name}.json のパース失敗: {ex}")
                            continue
                        outp = os.path.join(module_dir, f"{name}.json")
                        with open(outp, "w", encoding="utf-8") as jf:
                            json.dump(obj, jf, ensure_ascii=False, indent=2)

                # tfvars 本体へのパラメータ設定（policy はスキップ）
                for _, r in grp.iterrows():
                    if str(r.get("gen-tfvars-flag","")).strip().lower() != "true":
                        continue
                    arg = str(r.get("arguments","")).strip()
                    if arg == "policy":
                        continue
                    val = parse_value(r.get("value"))
                    set_nested_value(tfvars_data[top_key][name], arg, val)

    # 4) terraform.tfvars を modules/<env>/ に出力
    tfvars_dir = os.path.join(modules_root, env_value)
    os.makedirs(tfvars_dir, exist_ok=True)
    tfvars_path = os.path.join(tfvars_dir, "terraform.tfvars")
    with open(tfvars_path, "w", encoding="utf-8") as f:
        f.write(dict_to_hcl(tfvars_data, indent=0))

    # 5) terraform fmt (オプション)
    try:
        subprocess.run(["terraform", "fmt", tfvars_path], check=True)
    except Exception as e:
        print(f"terraform fmt に失敗: {e}")

if __name__ == "__main__":
    main()
