import os
import re
import pandas as pd
import subprocess
from openpyxl import load_workbook

# -------------------------------
# Excelのセル値を適切な型に変換する関数
def parse_value(val):
    """
    Excelから読み込んだ値を適切な型に変換する。
    例："20" -> 20, "gp3" -> "gp3", "True" -> True
    """
    if isinstance(val, str):
        val_strip = val.strip()
        if val_strip.lower() == "true":
            return True
        if val_strip.lower() == "false":
            return False
        try:
            return int(val_strip)
        except ValueError:
            try:
                return float(val_strip)
            except ValueError:
                return val_strip
    else:
        return val

# -------------------------------
# キー文字列からネストされた辞書へ値をセットする関数
def set_nested_value(d, key_str, value):
    """
    キー文字列（例："root_block_device[0].volume_size" や "tags.Name"）を分解し、
    辞書 d に対してネストした構造として value を設定する。
    """
    parts = key_str.split(".")
    current = d
    for i, part in enumerate(parts):
        m = re.match(r"(\w+)(?:\[(\d+)\])?", part)
        if not m:
            continue
        key = m.group(1)
        index = m.group(2)
        if i == len(parts) - 1:
            if index is not None:
                idx = int(index)
                if key not in current or not isinstance(current[key], list):
                    current[key] = []
                while len(current[key]) <= idx:
                    current[key].append(None)
                current[key][idx] = value
            else:
                current[key] = value
        else:
            if index is not None:
                idx = int(index)
                if key not in current or not isinstance(current[key], list):
                    current[key] = []
                while len(current[key]) <= idx:
                    current[key].append({})
                if current[key][idx] is None:
                    current[key][idx] = {}
                current = current[key][idx]
            else:
                if key not in current or not isinstance(current[key], dict):
                    current[key] = {}
                current = current[key]

# -------------------------------
# プリミティブ型の値をHCL形式へ変換する関数
def format_primitive(val):
    """
    文字列、数値、ブール値をHCL形式の文字列へ変換する。
    文字列はダブルクオートで囲み、ブールは小文字で出力する。
    """
    if isinstance(val, bool):
        return "true" if val else "false"
    elif isinstance(val, (int, float)):
        return str(val)
    else:
        return f'"{val}"'

# -------------------------------
# dict型をHCL形式の文字列へ変換する再帰関数
def dict_to_hcl(d, indent=0):
    """
    dict型をHCL形式の文字列へ変換する。
    """
    lines = []
    indent_str = "  " * indent
    for k, v in d.items():
        if isinstance(v, dict):
            lines.append(f'{indent_str}{k} = {{')
            lines.append(dict_to_hcl(v, indent + 1))
            lines.append(f'{indent_str}}}')
        elif isinstance(v, list):
            lines.append(f'{indent_str}{k} = {list_to_hcl(v, indent)}')
        else:
            lines.append(f'{indent_str}{k} = {format_primitive(v)}')
    return "\n".join(lines)

# -------------------------------
# list型をHCL形式の文字列へ変換する再帰関数
def list_to_hcl(lst, indent=0):
    """
    list型をHCL形式の文字列へ変換する。
    リストの要素がdictの場合は中括弧ブロックとして出力する。
    """
    indent_str = "  " * indent
    inner_indent = "  " * (indent + 1)
    lines = ["["]
    for item in lst:
        if isinstance(item, dict):
            lines.append(f'{inner_indent}' + "{")
            lines.append(dict_to_hcl(item, indent + 2))
            lines.append(f'{inner_indent}' + "},")
        elif isinstance(item, list):
            lines.append(f'{inner_indent}{list_to_hcl(item, indent + 1)},')
        else:
            lines.append(f'{inner_indent}{format_primitive(item)},')
    lines.append(f'{indent_str}]')
    return "\n".join(lines)

# -------------------------------
# openpyxlのワークブックのシートをDataFrameに変換する関数
def sheet_to_df(wb, sheet_name):
    """
    openpyxl のワークブックから指定シートを取得し、
    シートの最初の行をヘッダー、残りの行をデータとして DataFrame を返す。
    """
    ws = wb[sheet_name]
    data = ws.values
    try:
        header = next(data)
    except StopIteration:
        return pd.DataFrame()
    df = pd.DataFrame(data, columns=header)
    return df

# -------------------------------
# メイン処理
def main():
    # 入力フォルダ、出力フォルダの指定
    input_folder = "parameter"
    output_folder = "output_tfvars"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "output.tfvars")

    # sample.csv の存在確認と、シート名マッピングの読み込み
    mapping_sheet_names = {}
    csv_path = "sample.csv"
    if os.path.exists(csv_path):
        try:
            df_mapping = pd.read_csv(csv_path)
            for _, row in df_mapping.iterrows():
                excel_sheet_name = str(row["resource-name"]).strip()
                exact_name = str(row["exact-name"]).strip()
                mapping_sheet_names[excel_sheet_name] = exact_name
        except Exception as e:
            print("sample.csv の読み込みに失敗しました:", e)

    # tfvars 用データ格納用辞書
    tfvars_data = {}

    # 入力フォルダ内の全Excelファイルを処理
    for filename in os.listdir(input_folder):
        if not filename.lower().endswith(".xlsx"):
            continue
        excel_path = os.path.join(input_folder, filename)
        try:
            wb = load_workbook(excel_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"{excel_path} の読み込みに失敗しました: {e}")
            continue

        for sheet_name in wb.sheetnames:
            try:
                df = sheet_to_df(wb, sheet_name)
            except Exception as e:
                print(f"シート {sheet_name} を DataFrame に変換できませんでした: {e}")
                continue

            # sample.csv によるシート名のマッピングがあれば、正しいリソース種別名に修正
            actual_sheet_name = mapping_sheet_names.get(sheet_name, sheet_name)
            top_key = actual_sheet_name + "_list"
            if top_key not in tfvars_data:
                tfvars_data[top_key] = {}

            if "resource-name" not in df.columns:
                continue

            # 各リソース名ごとにグループ化して処理
            for resource_name in df["resource-name"].dropna().unique():
                resource_name = str(resource_name).strip()
                group = df[df["resource-name"].astype(str).str.strip() == resource_name]
                # まず、空の辞書を初期化（後でパラメータが設定される場合は上書きされる）
                if resource_name not in tfvars_data[top_key]:
                    tfvars_data[top_key][resource_name] = {}
                # gen-tfvars-flag が "true" の行が存在する場合のみ、値を設定する
                if group["gen-tfvars-flag"].astype(str).str.strip().str.lower().eq("true").any():
                    for _, row in group.iterrows():
                        flag = str(row.get("gen-tfvars-flag", "")).strip().lower()
                        if flag != "true":
                            continue
                        arguments = str(row.get("arguments")).strip()
                        raw_value = row.get("value")
                        value = parse_value(raw_value)
                        set_nested_value(tfvars_data[top_key][resource_name], arguments, value)
                # もし全ての行が "false" であれば、resource_name は空の辞書のままとなる

    # tfvars用 HCL 形式の文字列へ変換
    hcl_lines = dict_to_hcl(tfvars_data, indent=0)

    # 出力ファイルへ書き込み
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(hcl_lines)

    # 出力後、terraform fmt を実行する
    try:
        subprocess.run(["terraform", "fmt", output_file], check=True)
        print(f"tfvarsファイルを整形しました: {output_file}")
    except Exception as e:
        print("terraform fmt の実行に失敗しました。TerraformがPATHに含まれているか確認してください。")
        print(e)

if __name__ == "__main__":
    main()
