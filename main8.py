import os
import re
import json
import pandas as pd
import subprocess
from openpyxl import load_workbook

# ----------------------------------------
# Excelのセル値を適切な型に変換する関数
def parse_value(val):
    if isinstance(val, str):
        v = val.strip()
        if v.lower() == "true":   return True
        if v.lower() == "false":  return False
        try:    return int(v)
        except: 
            try:    return float(v)
            except: return v
    return val

# ----------------------------------------
# ネストキー文字列から辞書へ値をセットする関数
def set_nested_value(d, key_str, value):
    parts = key_str.split(".")
    cur = d
    for i, p in enumerate(parts):
        m = re.match(r"(\w+)(?:\[(\d+)\])?", p)
        if not m: continue
        key, idx = m.group(1), m.group(2)
        if i == len(parts)-1:
            if idx is not None:
                i0 = int(idx)
                cur.setdefault(key, [])
                while len(cur[key]) <= i0:
                    cur[key].append(None)
                cur[key][i0] = value
            else:
                cur[key] = value
        else:
            if idx is not None:
                i0 = int(idx)
                cur.setdefault(key, [])
                while len(cur[key]) <= i0:
                    cur[key].append({})
                if cur[key][i0] is None:
                    cur[key][i0] = {}
                cur = cur[key][i0]
            else:
                cur.setdefault(key, {})
                cur = cur[key]

# ----------------------------------------
# プリミティブ値→HCL文字列
def format_primitive(v):
    if isinstance(v, bool):    return "true" if v else "false"
    if isinstance(v, (int, float)): return str(v)
    return f'"{v}"'

# ----------------------------------------
# dict→HCL再帰整形
def dict_to_hcl(d, indent=0):
    ind = "  "*indent
    lines = []
    for k,v in d.items():
        if isinstance(v, dict):
            lines.append(f"{ind}{k} = {{")
            lines.append(dict_to_hcl(v, indent+1))
            lines.append(f"{ind}}}")
        elif isinstance(v, list):
            lines.append(f"{ind}{k} = {list_to_hcl(v, indent)}")
        else:
            lines.append(f"{ind}{k} = {format_primitive(v)}")
    return "\n".join(lines)

# ----------------------------------------
# list→HCL再帰整形
def list_to_hcl(lst, indent=0):
    ind = "  "*indent
    inner = "  "*(indent+1)
    lines = ["["]
    for it in lst:
        if isinstance(it, dict):
            lines.append(inner + "{")
            lines.append(dict_to_hcl(it, indent+2))
            lines.append(inner + "},")
        elif isinstance(it, list):
            lines.append(inner + list_to_hcl(it, indent+1) + ",")
        else:
            lines.append(inner + format_primitive(it) + ",")
    lines.append(ind + "]")
    return "\n".join(lines)

# ----------------------------------------
# openpyxlシート→DataFrame
def sheet_to_df(wb, name):
    ws = wb[name]
    vals = ws.values
    try:
        header = next(vals)
    except StopIteration:
        return pd.DataFrame()
    return pd.DataFrame(vals, columns=header)

# ----------------------------------------
def main():
    input_folder  = "parameter"
    output_folder = "output_tfvars"
    os.makedirs(output_folder, exist_ok=True)

    # ポリシーファイル用フォルダを別々に作成
    iam_folder = os.path.join(output_folder, "policies_iam")
    s3_folder  = os.path.join(output_folder, "policies_s3")
    os.makedirs(iam_folder, exist_ok=True)
    os.makedirs(s3_folder, exist_ok=True)

    # sample.csvでシート名マップ読込
    mapping = {}
    csv_path = "sample.csv"
    if os.path.exists(csv_path):
        dfm = pd.read_csv(csv_path)
        for _,r in dfm.iterrows():
            mapping[str(r["resource-name"])] = str(r["exact-name"])

    tfvars_data = {}

    for fn in os.listdir(input_folder):
        if not fn.lower().endswith(".xlsx"): continue
        path = os.path.join(input_folder, fn)
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            print(f"{fn} の読み込み失敗: {e}")
            continue

        for sheet in wb.sheetnames:
            df = sheet_to_df(wb, sheet)
            if df.empty or "resource-name" not in df.columns:
                continue

            actual = mapping.get(sheet, sheet)
            top = actual + "_list"
            tfvars_data.setdefault(top, {})

            for res in df["resource-name"].dropna().unique():
                rname = str(res).strip()
                group = df[df["resource-name"].astype(str).str.strip()==rname]
                tfvars_data[top].setdefault(rname, {})

                # policy用JSON出力
                # sheet 名が iam → iam_folder、s3 → s3_folder
                if group["arguments"].astype(str).str.strip().eq("policy").any():
                    # JSONとして正しくパースできるか試み
                    for _,row in group.iterrows():
                        if str(row["arguments"]).strip() != "policy":
                            continue
                        raw = row.get("value", "")
                        try:
                            obj = json.loads(raw)
                        except Exception as ex:
                            print(f"{rname}.json のパース失敗: {ex}")
                            continue
                        # 出力先フォルダを選択
                        if "iam" in sheet.lower():
                            outp = os.path.join(iam_folder, f"{rname}.json")
                        else:
                            outp = os.path.join(s3_folder,  f"{rname}.json")
                        with open(outp, "w", encoding="utf-8") as jf:
                            json.dump(obj, jf, ensure_ascii=False, indent=2)

                # tfvars本体には policy を含めずその他だけ
                for _,row in group.iterrows():
                    if str(row.get("gen-tfvars-flag","")).strip().lower() != "true":
                        continue
                    arg = str(row.get("arguments","")).strip()
                    # policy はスキップ
                    if arg == "policy":
                        continue
                    val = parse_value(row.get("value"))
                    set_nested_value(tfvars_data[top][rname], arg, val)

    # tfvars出力
    out_tfvars = os.path.join(output_folder, "output.tfvars")
    with open(out_tfvars, "w", encoding="utf-8") as f:
        f.write(dict_to_hcl(tfvars_data, indent=0))

    # terraform fmt
    try:
        subprocess.run(["terraform","fmt", out_tfvars], check=True)
    except:
        print("terraform fmt に失敗しました。")

if __name__=="__main__":
    main()
